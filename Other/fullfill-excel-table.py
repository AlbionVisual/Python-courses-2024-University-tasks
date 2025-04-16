import openpyxl
from datetime import date
from datetime import datetime
from datetime import timedelta
from sys import argv as entries

file_path = 'Weights.xlsx'

workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active

if len(entries) < 2: curr = date.today()
else: curr = datetime.strptime(entries[1], '%d.%m.%Y').date()

buf = 'Nothing'
i = worksheet.max_row - 1
for row in range(1, worksheet.max_row):
    if worksheet.cell(row=row, column=1).value is None:
        i = row - 1
        break

while buf[0] != 'w':
    buf = input(curr.strftime('%d.%m.%Y') + ': ')
    i += 1

    data = buf.split()
    if len(data) != 2:
        if buf[0] == 'w': break
        print('Try again')
        continue
    
    for j, val in enumerate(data):
        try:
            data[j] = float(val)
        except:
            data[j] = '-'

    worksheet.cell(row=i, column=1, value=curr)
    worksheet.cell(row=i, column=2, value=data[0])
    worksheet.cell(row=i, column=3, value=data[1])
    curr += timedelta(days=1)


workbook.save(file_path)